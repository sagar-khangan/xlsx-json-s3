import urllib.request
import xlrd
from openpyxl.workbook import Workbook as openpyxlWorkbook
from openpyxl import load_workbook
import json
import boto3

URL = 'https://www.iso20022.org/sites/default/files/ISO10383_MIC/ISO10383_MIC.xls/'
DEST_FILENAME = 'file.xlsx'
SHEET_NAME = "MICs List by CC"
DATA_FILE = 'data.json'
KEY = 'data.json'
BUCKET_NAME = 'san20'


def download_and_save():
	"""
		downloading xls file from url
		and converting it into xlsx
		workbook and saving in file

	"""
	try:
		xlsBook = xlrd.open_workbook(file_contents=urllib.request.urlopen(URL).read())
		workbook = openpyxlWorkbook()	
		for i in range(0, xlsBook.nsheets):
			xlsSheet = xlsBook.sheet_by_index(i)
			sheet = workbook.active if i == 0 else workbook.create_sheet()
			sheet.title = xlsSheet.name
			for row in range(0, xlsSheet.nrows):
				for col in range(0, xlsSheet.ncols):
			            sheet.cell(row=row + 1, column=col + 1).value = xlsSheet.cell_value(row, col)

		workbook.save(filename = DEST_FILENAME)
	except Exception as e:
		print("Error in downloading file: ", e)

def parse_file():
	"""

		parse file.xlsx and save data
		into data.json

	"""

	try:
		output = []
		workbook = load_workbook(DEST_FILENAME)
		worksheet = workbook[SHEET_NAME]
		headers = [ i.value for i in next(worksheet.iter_rows())]
		for row in worksheet.iter_rows():
			obj = {}
			for i in range(len(headers)):
				obj[headers[i]] = row[i].value	
			output.append(obj)
		with open(DATA_FILE, 'w') as fp:
			json.dump(output, fp, indent=4)
	except Exception as e:
		print("Error in parsing file: ", e)

def upload_s3():
	"""
	
		Uploads data.json to s3 bucket 
		with public access 
		url: https://s3.amazonaws.com/san20/data.json
		Credetntials set using aws configure in cli

	"""
	try:
		s3 = boto3.resource('s3')
		bucket = s3.Bucket(BUCKET_NAME)
		resp = bucket.upload_file(DATA_FILE, KEY, ExtraArgs={'ACL' : 'public-read'})
	except Exception as e:
		print("Error in uploading file to S3: ",e )

if __name__ == "__main__":
	print("downloading file")
	download_and_save()
	print("parsing xlsx file into json file")
	parse_file()
	print("uploading json file to S3 bucket")
	upload_s3()
	print("upload finished")
